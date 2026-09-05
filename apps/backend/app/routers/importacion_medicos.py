"""
Importación y exportación masiva de médicos para coordinadores.

Reglas de seguridad:
- Solo COORDINADOR con hospital asignado.
- El hospital SIEMPRE se deriva del coordinador autenticado (del token), nunca del
  Excel, el frontend, query params ni el body. Un coordinador no puede crear médicos
  en otro hospital.

Reutiliza la lógica de alta individual (`app.services.medico_service.crear_medico`)
y el servicio de correo (`app.services.email_service`).
"""

import io
import re
import unicodedata
from typing import Dict, List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.security import get_current_user
from app.db.db import get_db
from app.models.models import Hospital
from app.schemas.schemas import MedicoImportResult, MedicoImportErrorRow
from app.services import email_service
from app.services.medico_service import (
    crear_medico,
    generar_password_temporal,
    buscar_especialidades_por_nombre,
    MedicoValidationError,
)
from app.services.coordinador_service import obtener_coordinador_con_hospital

router = APIRouter()

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Encabezados esperados (clave canónica -> etiqueta visible)
COLUMNAS = {
    "nombre": "Nombre",
    "documento": "Cédula",
    "email": "Email",
    "telefono": "Teléfono",
    "especialidad": "Especialidad",
}
COLUMNAS_OBLIGATORIAS = ["nombre", "documento", "email"]

# Fila de ejemplo para la plantilla
EJEMPLO = {
    "nombre": "Juan Pérez",
    "documento": "1234567",
    "email": "juan@correo.com",
    "telefono": "0981123456",
    "especialidad": "Cardiología",
}

INSTRUCCIONES = [
    "Instrucciones para la importación masiva de médicos",
    "",
    "1. El archivo debe ser .xlsx.",
    "2. La primera fila debe contener los nombres de las columnas.",
    "3. No cambie los nombres de las columnas.",
    "4. No agregue una columna de hospital: los médicos se asocian automáticamente a su hospital.",
    "5. Cada fila representa un médico.",
    "6. Campos obligatorios: Nombre, Cédula, Email.",
    "7. El Email y la Cédula no pueden estar duplicados (ni en el sistema ni dentro del archivo).",
    "8. La Especialidad es opcional; si se indica, debe existir en el sistema y estar activa.",
    "   Puede indicar varias especialidades separadas por coma (ej: Cardiología, Clínica Médica).",
    "9. No incluya contraseñas: el sistema genera una contraseña temporal automáticamente",
    "   y la envía por correo a cada médico.",
]


def _normalizar(texto: Optional[str]) -> str:
    """Minúsculas, sin acentos y sin espacios sobrantes (para comparar encabezados)."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.strip().lower()


# Sinónimos aceptados por columna (normalizados)
SINONIMOS: Dict[str, str] = {
    "nombre": "nombre",
    "nombre completo": "nombre",
    "cedula": "documento",
    "documento": "documento",
    "documento de identidad": "documento",
    "ci": "documento",
    "email": "email",
    "correo": "email",
    "correo electronico": "email",
    "telefono": "telefono",
    "celular": "telefono",
    "especialidad": "especialidad",
    "especialidades": "especialidad",
}


def _celda_a_str(valor) -> str:
    """Convierte un valor de celda a string limpio (maneja None y números)."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip()


# ============================================================
# PLANTILLA
# ============================================================

@router.get("/plantilla")
def descargar_plantilla(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Descarga una plantilla .xlsx con las columnas esperadas y una fila de ejemplo."""
    obtener_coordinador_con_hospital(db, current_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "Médicos"

    encabezados = [COLUMNAS[k] for k in COLUMNAS]
    ws.append(encabezados)
    ws.append([EJEMPLO[k] for k in COLUMNAS])

    # Ancho de columnas para legibilidad
    for idx, _ in enumerate(encabezados, start=1):
        ws.column_dimensions[chr(64 + idx)].width = 24

    ws_info = wb.create_sheet("Instrucciones")
    for linea in INSTRUCCIONES:
        ws_info.append([linea])
    ws_info.column_dimensions["A"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="plantilla_medicos.xlsx"'},
    )


# ============================================================
# IMPORTACIÓN
# ============================================================

@router.post("/importar", response_model=MedicoImportResult)
def importar_medicos(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Importa médicos desde un archivo .xlsx y los asocia al hospital del coordinador.
    Genera una contraseña temporal para cada uno y envía un correo de bienvenida.
    Las filas con error no detienen el proceso: se crean las válidas y se reportan las inválidas.
    """
    coordinador = obtener_coordinador_con_hospital(db, current_user)
    hospital: Hospital = coordinador.hospital

    # Validar extensión
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe tener formato .xlsx",
        )

    # Abrir el Excel
    try:
        contenido = file.file.read()
        wb = load_workbook(filename=io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se pudo leer el archivo. Asegúrese de que sea un .xlsx válido.",
        )

    filas = ws.iter_rows(values_only=True)

    # Leer encabezados
    try:
        encabezados = next(filas)
    except StopIteration:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo está vacío.",
        )

    # Mapear encabezados -> índice de columna (por clave canónica)
    columna_por_clave: Dict[str, int] = {}
    for idx, encabezado in enumerate(encabezados or []):
        clave = SINONIMOS.get(_normalizar(encabezado))
        if clave and clave not in columna_por_clave:
            columna_por_clave[clave] = idx

    # Validar columnas obligatorias ANTES de procesar filas
    faltantes = [COLUMNAS[k] for k in COLUMNAS_OBLIGATORIAS if k not in columna_por_clave]
    if faltantes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo no contiene las columnas requeridas: {', '.join(faltantes)}.",
        )

    def _valor(fila, clave: str) -> str:
        idx = columna_por_clave.get(clave)
        if idx is None or idx >= len(fila):
            return ""
        return _celda_a_str(fila[idx])

    procesados = 0
    creados = 0
    con_error = 0
    correos_enviados = 0
    correos_con_error = 0
    errores: List[MedicoImportErrorRow] = []

    emails_vistos: set = set()
    documentos_vistos: set = set()

    # Empezamos a numerar en 2 (fila 1 = encabezados)
    for offset, fila in enumerate(filas, start=2):
        # Saltar filas completamente vacías
        if fila is None or all(_celda_a_str(c) == "" for c in fila):
            continue

        nombre = _valor(fila, "nombre")
        documento = _valor(fila, "documento")
        email = _valor(fila, "email")
        telefono = _valor(fila, "telefono") or None
        especialidad_raw = _valor(fila, "especialidad")

        procesados += 1
        etiqueta = nombre or email or documento or "(sin nombre)"

        # Validaciones de fila
        campos_faltantes = []
        if not nombre:
            campos_faltantes.append("Nombre")
        if not documento:
            campos_faltantes.append("Cédula")
        if not email:
            campos_faltantes.append("Email")
        if campos_faltantes:
            con_error += 1
            errores.append(MedicoImportErrorRow(
                fila=offset, medico=etiqueta,
                resultado=f"Faltan campos obligatorios: {', '.join(campos_faltantes)}",
            ))
            continue

        if not EMAIL_REGEX.match(email):
            con_error += 1
            errores.append(MedicoImportErrorRow(fila=offset, medico=etiqueta, resultado="Email inválido"))
            continue

        email_norm = email.lower()
        if email_norm in emails_vistos:
            con_error += 1
            errores.append(MedicoImportErrorRow(
                fila=offset, medico=etiqueta, resultado="Email duplicado dentro del archivo",
            ))
            continue
        if documento in documentos_vistos:
            con_error += 1
            errores.append(MedicoImportErrorRow(
                fila=offset, medico=etiqueta, resultado="Cédula duplicada dentro del archivo",
            ))
            continue

        # Resolver especialidades (opcional; una o varias separadas por coma)
        try:
            nombres_esp = [n for n in (especialidad_raw.split(",") if especialidad_raw else []) if n.strip()]
            especialidades = buscar_especialidades_por_nombre(db, nombres_esp)
        except MedicoValidationError as e:
            con_error += 1
            errores.append(MedicoImportErrorRow(fila=offset, medico=etiqueta, resultado=str(e)))
            continue

        # Crear médico (commit por fila para aislar errores)
        password_temporal = generar_password_temporal()
        try:
            crear_medico(
                db,
                documento=documento,
                nombre=nombre,
                email=email,
                telefono=telefono,
                password=password_temporal,
                especialidades=especialidades,
                hospitales=[hospital],
                debe_cambiar_password=True,
                commit=True,
            )
        except MedicoValidationError as e:
            db.rollback()
            con_error += 1
            errores.append(MedicoImportErrorRow(fila=offset, medico=etiqueta, resultado=str(e)))
            continue
        except Exception as e:  # noqa: BLE001
            db.rollback()
            con_error += 1
            errores.append(MedicoImportErrorRow(fila=offset, medico=etiqueta, resultado=f"Error al crear: {e}"))
            continue

        creados += 1
        emails_vistos.add(email_norm)
        documentos_vistos.add(documento)

        # Enviar correo de bienvenida (un fallo NO elimina al médico ya creado)
        try:
            email_service.enviar_bienvenida_medico(
                email=email,
                nombre=nombre,
                hospital_nombre=hospital.nombre,
                password_temporal=password_temporal,
            )
            correos_enviados += 1
        except email_service.EmailNoConfiguradoError:
            correos_con_error += 1
            errores.append(MedicoImportErrorRow(
                fila=offset, medico=etiqueta,
                resultado="Médico creado, pero el correo no se envió (SMTP no configurado)",
            ))
        except Exception as e:  # noqa: BLE001 - cualquier fallo de envío
            correos_con_error += 1
            errores.append(MedicoImportErrorRow(
                fila=offset, medico=etiqueta,
                resultado=f"Médico creado, pero falló el envío del correo: {e}",
            ))

    return MedicoImportResult(
        hospital=hospital.nombre,
        procesados=procesados,
        creados=creados,
        con_error=con_error,
        correos_enviados=correos_enviados,
        correos_con_error=correos_con_error,
        errores=errores,
    )


# ============================================================
# EXPORTACIÓN
# ============================================================

@router.get("/exportar")
def exportar_medicos(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Exporta a .xlsx los médicos del hospital del coordinador (sin datos sensibles de auth)."""
    coordinador = obtener_coordinador_con_hospital(db, current_user)
    hospital: Hospital = coordinador.hospital

    wb = Workbook()
    ws = wb.active
    ws.title = "Médicos"

    encabezados = ["Nombre", "Documento", "Email", "Teléfono", "Especialidades", "Hospital"]
    ws.append(encabezados)

    for medico in hospital.medicos:
        especialidades = ", ".join(e.nombre for e in medico.especialidades)
        ws.append([
            medico.nombre,
            medico.documento,
            medico.email,
            medico.telefono or "",
            especialidades,
            hospital.nombre,
        ])

    for idx, _ in enumerate(encabezados, start=1):
        ws.column_dimensions[chr(64 + idx)].width = 24

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="medicos_hospital.xlsx"'},
    )
