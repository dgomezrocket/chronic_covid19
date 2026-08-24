from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook, load_workbook
from app.db.db import get_db
from app.models.models import Hospital, Paciente
from app.schemas.schemas import (
    HospitalCreate,
    HospitalUpdate,
    HospitalOut,
    HospitalesCercanosResponse,
    HospitalConDistanciaOut,
    HospitalImportResult,
    HospitalImportErrorRow,
)
from app.core.security import get_current_user
import csv
import io
import math
import unicodedata

router = APIRouter()

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# ============================================================
# IMPORTACIÓN / EXPORTACIÓN EN EXCEL
# Formato compartido por la plantilla, la importación y la exportación.
# ============================================================

# Clave canónica del campo -> etiqueta visible en el Excel
COLUMNAS: Dict[str, str] = {
    "nombre": "Nombre",
    "codigo": "Código",
    "departamento": "Departamento",
    "ciudad": "Ciudad",
    "barrio": "Barrio",
    "direccion": "Dirección",
    "telefono": "Teléfono",
    "latitud": "Latitud",
    "longitud": "Longitud",
}

# Todos los datos del hospital son obligatorios
CAMPOS_TEXTO = ["nombre", "codigo", "departamento", "ciudad", "barrio", "direccion", "telefono"]
CAMPOS_NUMERICOS = ["latitud", "longitud"]
CAMPOS_OBLIGATORIOS = CAMPOS_TEXTO + CAMPOS_NUMERICOS

EJEMPLO: Dict[str, str] = {
    "nombre": "Hospital General de Luque",
    "codigo": "HGL-001",
    "departamento": "Central",
    "ciudad": "Luque",
    "barrio": "Centro",
    "direccion": "Av. Humaitá 123",
    "telefono": "021123456",
    "latitud": "-25.2678",
    "longitud": "-57.4872",
}

INSTRUCCIONES = [
    "Instrucciones para la importación masiva de hospitales",
    "",
    "1. El archivo debe ser .xlsx.",
    "2. No modifique los nombres de las columnas.",
    "3. Cada fila representa un hospital.",
    "4. Todos los campos son obligatorios.",
    "5. El Código debe ser único (en el sistema y dentro del archivo).",
    "6. La Latitud debe ser un número entre -90 y 90.",
    "7. La Longitud debe ser un número entre -180 y 180.",
    "8. No agregue columnas innecesarias.",
    "9. No deje celdas obligatorias vacías.",
]

# Sinónimos aceptados por columna (normalizados: sin acentos, en minúsculas).
# Se conservan los nombres antiguos del CSV (provincia/distrito) por compatibilidad.
SINONIMOS: Dict[str, str] = {
    "nombre": "nombre",
    "nombre del hospital": "nombre",
    "hospital": "nombre",
    "codigo": "codigo",
    "departamento": "departamento",
    "provincia": "departamento",
    "ciudad": "ciudad",
    "distrito": "ciudad",
    "barrio": "barrio",
    "direccion": "direccion",
    "telefono": "telefono",
    "latitud": "latitud",
    "lat": "latitud",
    "longitud": "longitud",
    "lon": "longitud",
    "lng": "longitud",
}


def _require_admin(current_user: dict, accion: str) -> None:
    """Valida que el usuario autenticado tenga rol admin."""
    if current_user.get("rol") != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail=f"Solo los administradores pueden {accion}",
        )


def _normalizar(texto: Optional[str]) -> str:
    """Minúsculas, sin acentos y sin espacios sobrantes (para comparar encabezados)."""
    if texto is None:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.strip().lower()


def _celda_a_str(valor) -> str:
    """Convierte un valor de celda a string limpio (maneja None y números)."""
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    return str(valor).strip()


def _parsear_coordenada(valor: str, minimo: float, maximo: float) -> float:
    """Convierte a float y valida el rango. Lanza ValueError con un mensaje legible."""
    try:
        numero = float(valor.replace(",", "."))
    except (TypeError, ValueError):
        raise ValueError("inválida")
    if not math.isfinite(numero):
        raise ValueError("inválida")
    if not (minimo <= numero <= maximo):
        raise ValueError(f"fuera del rango permitido ({minimo} a {maximo})")
    return numero


def _validar_hospital_completo(datos: dict) -> List[str]:
    """
    Devuelve las etiquetas de los campos obligatorios que faltan.
    Se usa al editar para garantizar que el hospital quede completo.
    """
    faltantes = []
    for clave in CAMPOS_TEXTO:
        valor = datos.get(clave)
        if valor is None or not str(valor).strip():
            faltantes.append(COLUMNAS[clave])
    for clave in CAMPOS_NUMERICOS:
        if datos.get(clave) is None:
            faltantes.append(COLUMNAS[clave])
    return faltantes


def _autoajustar_columnas(ws, encabezados: List[str], ancho: int = 24) -> None:
    """Aplica un ancho fijo legible a las columnas del Excel."""
    for idx, _ in enumerate(encabezados, start=1):
        ws.column_dimensions[chr(64 + idx)].width = ancho


def _distancia_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Calcula la distancia en km entre dos coordenadas usando la fórmula de Haversine."""
    R = 6371  # Radio de la Tierra en km
    d_lat = math.radians(lat2 - lat1)
    d_lon = math.radians(lon2 - lon1)
    a = (
        math.sin(d_lat / 2) ** 2
        + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(d_lon / 2) ** 2
    )
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return R * c


@router.get("/", response_model=List[HospitalOut])
def get_all_hospitales(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    nombre: Optional[str] = None,
    departamento: Optional[str] = None,
    ciudad: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Obtiene todos los hospitales con filtros opcionales (público)"""
    query = db.query(Hospital)

    # Aplicar filtros si existen
    if nombre:
        query = query.filter(Hospital.nombre.ilike(f"%{nombre}%"))
    if departamento:
        query = query.filter(Hospital.departamento.ilike(f"%{departamento}%"))
    if ciudad:
        query = query.filter(Hospital.ciudad.ilike(f"%{ciudad}%"))

    hospitales = query.offset(skip).limit(limit).all()
    return hospitales


@router.get("/mis-cercanos", response_model=HospitalesCercanosResponse)
def get_mis_hospitales_cercanos(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Devuelve los hospitales del sistema ordenados del más cercano al más lejano
    respecto de la ubicación registrada del paciente autenticado.

    La ubicación se obtiene siempre del paciente del token (nunca de un ID enviado
    por el cliente). Solo accesible para pacientes.
    """
    # ✅ Validar que sea paciente
    if current_user["rol"] != "paciente":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo los pacientes pueden buscar hospitales cercanos"
        )

    paciente = db.query(Paciente).filter(Paciente.id == current_user["id"]).first()
    if not paciente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Paciente no encontrado"
        )

    # Sin ubicación registrada: no es un error, devolvemos la bandera para el frontend
    if paciente.latitud is None or paciente.longitud is None:
        return HospitalesCercanosResponse(
            tiene_ubicacion=False,
            latitud=None,
            longitud=None,
            hospitales=[]
        )

    hospitales = db.query(Hospital).all()
    hospitales_con_distancia = []
    for h in hospitales:
        if h.latitud is None or h.longitud is None:
            continue
        distancia = _distancia_km(paciente.latitud, paciente.longitud, h.latitud, h.longitud)
        hospitales_con_distancia.append(
            HospitalConDistanciaOut.model_validate(
                {**HospitalOut.model_validate(h).model_dump(), "distancia_km": round(distancia, 2)}
            )
        )

    # Ordenar del más cercano al más lejano
    hospitales_con_distancia.sort(key=lambda x: x.distancia_km if x.distancia_km is not None else float("inf"))

    return HospitalesCercanosResponse(
        tiene_ubicacion=True,
        latitud=paciente.latitud,
        longitud=paciente.longitud,
        hospitales=hospitales_con_distancia
    )


@router.get("/nearby", response_model=List[HospitalOut])
def get_hospitales_cercanos(
    lat: float = Query(...),
    lon: float = Query(...),
    radio: float = Query(5.0),
    db: Session = Depends(get_db)
):
    """Obtiene hospitales cercanos a una ubicación (público)"""
    hospitales = db.query(Hospital).all()
    cercanos = []

    for h in hospitales:
        if h.latitud and h.longitud:
            # Cálculo simple de distancia (no es exacto pero funciona para distancias cortas)
            distancia = ((h.latitud - lat)**2 + (h.longitud - lon)**2)**0.5
            if distancia <= radio:
                cercanos.append(h)

    return cercanos


# ============================================================
# PLANTILLA EXCEL
# Declarada antes de /{hospital_id} para que FastAPI no interprete
# "plantilla" como un hospital_id.
# ============================================================

@router.get("/plantilla")
def descargar_plantilla_hospitales(
    current_user: dict = Depends(get_current_user),
):
    """Descarga una plantilla .xlsx con las columnas esperadas y una fila de ejemplo (solo admin)."""
    _require_admin(current_user, "descargar la plantilla de hospitales")

    wb = Workbook()
    ws = wb.active
    ws.title = "Hospitales"

    encabezados = [COLUMNAS[k] for k in COLUMNAS]
    ws.append(encabezados)
    ws.append([EJEMPLO[k] for k in COLUMNAS])
    _autoajustar_columnas(ws, encabezados)

    ws_info = wb.create_sheet("Instrucciones")
    for linea in INSTRUCCIONES:
        ws_info.append([linea])
    ws_info.column_dimensions["A"].width = 90

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="plantilla_hospitales.xlsx"'},
    )


# ============================================================
# EXPORTACIÓN EXCEL
# Declarada antes de /{hospital_id} (ver nota de la plantilla).
# ============================================================

@router.get("/exportar")
def exportar_hospitales(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Exporta a .xlsx todos los hospitales del sistema (solo admin).

    Se consultan siempre todos los hospitales desde la base, sin la paginación de
    GET /hospitales/, y no se incluyen IDs internos.
    """
    _require_admin(current_user, "exportar hospitales")

    wb = Workbook()
    ws = wb.active
    ws.title = "Hospitales"

    encabezados = [COLUMNAS[k] for k in COLUMNAS]
    ws.append(encabezados)

    hospitales = db.query(Hospital).order_by(Hospital.nombre).all()
    for h in hospitales:
        ws.append([
            h.nombre or "",
            h.codigo or "",
            h.departamento or "",
            h.ciudad or "",
            h.barrio or "",
            h.direccion or "",
            h.telefono or "",
            h.latitud if h.latitud is not None else "",
            h.longitud if h.longitud is not None else "",
        ])

    _autoajustar_columnas(ws, encabezados)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="hospitales.xlsx"'},
    )


@router.get("/{hospital_id}", response_model=HospitalOut)
def get_hospital_by_id(
    hospital_id: int,
    db: Session = Depends(get_db)
):
    """Obtiene un hospital por ID (público)"""
    hospital = db.query(Hospital).filter(Hospital.id == hospital_id).first()

    if not hospital:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Hospital no encontrado"
        )

    return hospital


@router.post("/", response_model=HospitalOut, status_code=status.HTTP_201_CREATED)
def create_hospital(
    hospital: HospitalCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Crea un nuevo hospital (solo admin).

    Todos los datos del hospital son obligatorios: HospitalCreate los valida
    (incluido el rango de latitud/longitud) antes de llegar aquí.
    """
    # ✅ Validar que sea admin
    _require_admin(current_user, "crear hospitales")

    # Verificar si ya existe un hospital con ese código
    existing = db.query(Hospital).filter(Hospital.codigo == hospital.codigo).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Ya existe un hospital con ese código"
        )

    # Crear el hospital
    nuevo_hospital = Hospital(
        nombre=hospital.nombre,
        codigo=hospital.codigo,
        ciudad=hospital.ciudad,
        departamento=hospital.departamento,
        barrio=hospital.barrio,
        direccion=hospital.direccion,
        telefono=hospital.telefono,
        latitud=hospital.latitud,
        longitud=hospital.longitud
    )

    db.add(nuevo_hospital)
    db.commit()
    db.refresh(nuevo_hospital)

    return nuevo_hospital


@router.put("/{hospital_id}", response_model=HospitalOut)
def update_hospital(
    hospital_id: int,
    hospital_update: HospitalUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Actualiza un hospital (solo admin).

    La edición sigue siendo parcial, pero al guardar el hospital debe quedar con
    todos los datos obligatorios completos. Los registros históricos que no se
    editan no se modifican.
    """
    # ✅ Validar que sea admin
    _require_admin(current_user, "actualizar hospitales")

    hospital = db.query(Hospital).filter(Hospital.id == hospital_id).first()

    if not hospital:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Hospital no encontrado"
        )

    # Verificar código duplicado si se está cambiando
    update_data = hospital_update.model_dump(exclude_unset=True)
    if "codigo" in update_data and update_data["codigo"] != hospital.codigo:
        existing = db.query(Hospital).filter(
            Hospital.codigo == update_data["codigo"]
        ).first()
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un hospital con ese código"
            )

    # El hospital debe quedar con todos los datos obligatorios completos
    estado_final = {
        clave: update_data.get(clave, getattr(hospital, clave))
        for clave in CAMPOS_OBLIGATORIOS
    }
    faltantes = _validar_hospital_completo(estado_final)
    if faltantes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan campos obligatorios: {', '.join(faltantes)}"
        )

    # Actualizar campos
    for field, value in update_data.items():
        setattr(hospital, field, value)

    db.commit()
    db.refresh(hospital)

    return hospital


@router.delete("/{hospital_id}", status_code=status.HTTP_200_OK)
def delete_hospital(
    hospital_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """Elimina un hospital (solo admin)"""
    # ✅ Validar que sea admin
    _require_admin(current_user, "eliminar hospitales")

    hospital = db.query(Hospital).filter(Hospital.id == hospital_id).first()

    if not hospital:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Hospital no encontrado"
        )

    # Verificar si hay médicos asignados a este hospital
    if hospital.medicos and len(hospital.medicos) > 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"No se puede eliminar el hospital porque tiene {len(hospital.medicos)} médico(s) asignado(s)"
        )

    db.delete(hospital)
    db.commit()

    return {"message": "Hospital eliminado exitosamente", "id": hospital_id}


# ============================================================
# IMPORTACIÓN MASIVA
# ============================================================

def _leer_encabezados_y_filas(file: UploadFile) -> Tuple[List[str], List[List[str]]]:
    """
    Lee el archivo subido y devuelve (encabezados, filas) como texto.

    Formato principal: .xlsx (openpyxl). Se conserva el soporte de .csv del
    endpoint anterior para no romper el comportamiento previo.
    """
    filename = (file.filename or "").lower()
    contenido = file.file.read()

    if not contenido:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo está vacío.",
        )

    if filename.endswith(".csv"):
        texto = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                texto = contenido.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if texto is None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se pudo leer el archivo CSV (codificación no soportada).",
            )
        lector = csv.reader(io.StringIO(texto))
        try:
            encabezados = next(lector)
        except StopIteration:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El archivo está vacío.",
            )
        filas = [[_celda_a_str(c) for c in fila] for fila in lector]
        return [_celda_a_str(c) for c in encabezados], filas

    if not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo debe tener formato .xlsx (también se acepta .csv por compatibilidad).",
        )

    try:
        wb = load_workbook(filename=io.BytesIO(contenido), read_only=True, data_only=True)
        ws = wb.active
        filas_iter = ws.iter_rows(values_only=True)
        encabezados = next(filas_iter)
        filas = [[_celda_a_str(c) for c in (fila or [])] for fila in filas_iter]
    except StopIteration:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El archivo está vacío.",
        )
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No se pudo leer el archivo. Asegúrese de que sea un .xlsx válido.",
        )

    return [_celda_a_str(c) for c in (encabezados or [])], filas


@router.post("/import", response_model=HospitalImportResult)
def importar_hospitales(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    """
    Importa hospitales desde un archivo Excel .xlsx (solo admin).

    Se mantiene el soporte de .csv del comportamiento anterior. Todos los campos
    del hospital son obligatorios y el código debe ser único: las filas inválidas
    se reportan pero no detienen el proceso. No hay actualización automática
    (upsert) de hospitales existentes.
    """
    # ✅ Validar que sea admin
    _require_admin(current_user, "importar hospitales")

    encabezados, filas = _leer_encabezados_y_filas(file)

    # Mapear encabezados -> índice de columna (por clave canónica)
    columna_por_clave: Dict[str, int] = {}
    for idx, encabezado in enumerate(encabezados):
        clave = SINONIMOS.get(_normalizar(encabezado))
        if clave and clave not in columna_por_clave:
            columna_por_clave[clave] = idx

    # Validar columnas obligatorias ANTES de procesar filas
    faltantes = [COLUMNAS[k] for k in CAMPOS_OBLIGATORIOS if k not in columna_por_clave]
    if faltantes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El archivo no contiene las columnas requeridas: {', '.join(faltantes)}.",
        )

    def _valor(fila: List[str], clave: str) -> str:
        idx = columna_por_clave.get(clave)
        if idx is None or idx >= len(fila):
            return ""
        return fila[idx]

    procesados = 0
    importados = 0
    con_error = 0
    errores: List[HospitalImportErrorRow] = []

    codigos_vistos: set = set()

    # Empezamos a numerar en 2 (fila 1 = encabezados)
    for offset, fila in enumerate(filas, start=2):
        # Saltar filas completamente vacías
        if not fila or all(c == "" for c in fila):
            continue

        valores = {clave: _valor(fila, clave) for clave in CAMPOS_OBLIGATORIOS}
        etiqueta = valores["nombre"] or valores["codigo"] or "(sin nombre)"
        procesados += 1

        def _agregar_error(motivo: str, _fila=offset, _etiqueta=etiqueta) -> None:
            errores.append(HospitalImportErrorRow(fila=_fila, hospital=_etiqueta, resultado=motivo))

        # 1) Campos obligatorios presentes (no se acepta cadena vacía)
        campos_faltantes = [COLUMNAS[k] for k in CAMPOS_OBLIGATORIOS if not valores[k]]
        if campos_faltantes:
            con_error += 1
            _agregar_error(f"Faltan campos obligatorios: {', '.join(campos_faltantes)}")
            continue

        # 2) Coordenadas numéricas y dentro de rango
        try:
            latitud = _parsear_coordenada(valores["latitud"], -90, 90)
        except ValueError as e:
            con_error += 1
            _agregar_error(f"Latitud {e}")
            continue
        try:
            longitud = _parsear_coordenada(valores["longitud"], -180, 180)
        except ValueError as e:
            con_error += 1
            _agregar_error(f"Longitud {e}")
            continue

        # 3) Código único dentro del archivo y en el sistema (sin upsert)
        codigo = valores["codigo"]
        if codigo in codigos_vistos:
            con_error += 1
            _agregar_error(f"Código {codigo} duplicado dentro del archivo")
            continue
        if db.query(Hospital).filter(Hospital.codigo == codigo).first():
            con_error += 1
            _agregar_error(f"Código {codigo} ya registrado")
            continue

        # 4) Alta (commit por fila para aislar errores)
        try:
            hospital = Hospital(
                nombre=valores["nombre"],
                codigo=codigo,
                departamento=valores["departamento"],
                ciudad=valores["ciudad"],
                barrio=valores["barrio"],
                direccion=valores["direccion"],
                telefono=valores["telefono"],
                latitud=latitud,
                longitud=longitud,
            )
            db.add(hospital)
            db.commit()
        except Exception as e:  # noqa: BLE001
            db.rollback()
            con_error += 1
            _agregar_error(f"Error al crear: {e}")
            continue

        importados += 1
        codigos_vistos.add(codigo)

    return HospitalImportResult(
        procesados=procesados,
        importados=importados,
        con_error=con_error,
        errores=errores,
        total_errores=len(errores),
    )
